import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00, FORMAT_DATE_YYYYMMDD2
import os
from datetime import datetime

def gerar_relatorio_mensal_personalizado(registros, funcionarios, projetos, mes_ano, caminho_saida):
    """
    Gera um relatório mensal personalizado no formato do template fornecido pelo usuário.
    
    Args:
        registros: Lista de registros de horas
        funcionarios: Lista de funcionários
        projetos: Lista de projetos
        mes_ano: Mês/ano de referência no formato "YYYY-MM"
        caminho_saida: Caminho para salvar o arquivo Excel
    
    Returns:
        Caminho do arquivo Excel gerado
    """
    # Criar um novo workbook e selecionar a planilha ativa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha1"
    
    # Definir estilos
    header_font = Font(name='Calibri', size=11, bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    left_border = Border(
        left=Side(style='thin'),
        right=Side(style=None),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    right_border = Border(
        left=Side(style=None),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    middle_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definir cabeçalhos na linha 3
    headers = [
        'ID', 'Id Colaborador', 'Nome Colaborador', 'Mês', 'Horas Mês',
        'GP', 'Horas Trabalhadas', 'Proporção de Hora', 'Valor por GP',
        'GP 9014', 'Observação GP 9010', 'Observação GP 9021'
    ]
    
    # Adicionar cabeçalhos começando na coluna C
    for i, header in enumerate(headers):
        col = i + 3  # Começando na coluna C (índice 3)
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        
        # Aplicar bordas apropriadas
        if i == 0:  # Primeira coluna (ID)
            cell.border = thin_border
        elif i == len(headers) - 1:  # Última coluna
            cell.border = thin_border
        else:  # Colunas do meio
            cell.border = middle_border
    
    # Preparar dados para o relatório
    row_index = 4  # Começar na linha 4 (após os cabeçalhos)
    id_counter = 1
    
    # Dicionário para agrupar registros por funcionário
    registros_por_funcionario = {}
    for registro in registros:
        funcionario_id = registro.funcionario_id
        if funcionario_id not in registros_por_funcionario:
            registros_por_funcionario[funcionario_id] = []
        registros_por_funcionario[funcionario_id].append(registro)
    
    # Projetos especiais - usar get_projeto_by_id para evitar erros
    projeto_9014 = None
    projeto_9010 = None
    projeto_9021 = None
    
    for p in projetos:
        if p.id == 9014:
            projeto_9014 = p
        elif p.id == 9010:
            projeto_9010 = p
        elif p.id == 9021:
            projeto_9021 = p
    
    # Preencher dados
    for funcionario_id, registros_func in registros_por_funcionario.items():
        funcionario = None
        for f in funcionarios:
            if f.id == funcionario_id:
                funcionario = f
                break
                
        if not funcionario:
            continue
        
        # Calcular horas totais do funcionário no mês
        horas_totais = sum(r.horas_trabalhadas for r in registros_func)
        
        # Agrupar registros por projeto
        registros_por_projeto = {}
        for registro in registros_func:
            projeto_id = registro.projeto_id
            if projeto_id not in registros_por_projeto:
                registros_por_projeto[projeto_id] = []
            registros_por_projeto[projeto_id].append(registro)
        
        # Adicionar uma linha para cada projeto do funcionário
        for projeto_id, registros_proj in registros_por_projeto.items():
            projeto = None
            for p in projetos:
                if p.id == projeto_id:
                    projeto = p
                    break
                    
            if not projeto:
                continue
            
            # Calcular horas trabalhadas neste projeto
            horas_projeto = sum(r.horas_trabalhadas for r in registros_proj)
            
            # Calcular proporção de horas
            proporcao = horas_projeto / horas_totais if horas_totais > 0 else 0
            
            # Adicionar dados à planilha
            ws.cell(row=row_index, column=3, value=id_counter)  # ID
            ws.cell(row=row_index, column=4, value=funcionario_id)  # Id Colaborador
            ws.cell(row=row_index, column=5, value=funcionario.nome)  # Nome Colaborador
            ws.cell(row=row_index, column=6, value=mes_ano)  # Mês
            ws.cell(row=row_index, column=7, value=horas_totais)  # Horas Mês
            ws.cell(row=row_index, column=8, value=projeto_id)  # GP (código do projeto)
            ws.cell(row=row_index, column=9, value=horas_projeto)  # Horas Trabalhadas
            ws.cell(row=row_index, column=10, value=proporcao)  # Proporção de Hora
            
            # Formatação de números
            ws.cell(row=row_index, column=7).number_format = FORMAT_NUMBER_00
            ws.cell(row=row_index, column=9).number_format = FORMAT_NUMBER_00
            ws.cell(row=row_index, column=10).number_format = '0.00%'
            
            # Observações para projetos especiais
            if projeto_id == 9014:  # GP 9014 (Propostas)
                # Buscar observações nos registros deste projeto
                observacoes = [r.observacao for r in registros_proj if hasattr(r, 'observacao') and r.observacao]
                ws.cell(row=row_index, column=11, value=", ".join(observacoes) if observacoes else "")
            
            if projeto_id == 9010:  # GP 9010 (Atividades Internas)
                observacoes = [r.observacao for r in registros_proj if hasattr(r, 'observacao') and r.observacao]
                ws.cell(row=row_index, column=12, value=", ".join(observacoes) if observacoes else "")
            
            if projeto_id == 9021:  # GP 9021 (Férias e Recessos)
                observacoes = [r.observacao for r in registros_proj if hasattr(r, 'observacao') and r.observacao]
                ws.cell(row=row_index, column=13, value=", ".join(observacoes) if observacoes else "")
            
            # Aplicar bordas às células de dados
            for col in range(3, 15):  # Colunas C a N
                cell = ws.cell(row=row_index, column=col)
                if col == 3:  # Primeira coluna (C)
                    cell.border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
                elif col == 14:  # Última coluna (N)
                    cell.border = Border(right=Side(style='thin'), bottom=Side(style='thin'))
                else:
                    cell.border = Border(bottom=Side(style='thin'))
            
            row_index += 1
            id_counter += 1
    
    # Ajustar largura das colunas
    for col in range(3, 15):  # Colunas C a N
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
    
    # Coluna de nome do colaborador mais larga
    ws.column_dimensions['E'].width = 25
    
    # Salvar o arquivo
    try:
        wb.save(caminho_saida)
        return caminho_saida
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        # Tentar salvar em um local alternativo
        alt_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), f"controle_horas_alt_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(alt_path)
        return alt_path

# Função para adaptar o código de exportação existente
def adaptar_exportacao_relatorio_mensal(db, funcionario_id=None, projeto_id=None, mes_ano=None):
    """
    Adapta o código de exportação existente para usar o novo formato personalizado.
    
    Args:
        db: Instância do banco de dados
        funcionario_id: ID do funcionário para filtrar (opcional)
        projeto_id: ID do projeto para filtrar (opcional)
        mes_ano: Mês/ano para filtrar no formato "YYYY-MM" (opcional)
    
    Returns:
        Caminho do arquivo Excel gerado
    """
    try:
        # Obter os registros filtrados
        registros = db.listar_registros_horas(
            funcionario_id=funcionario_id,
            projeto_id=projeto_id,
            mes_ano=mes_ano
        )
        
        # Obter listas de funcionários e projetos
        funcionarios = db.listar_funcionarios()
        projetos = db.listar_projetos()
        
        # Verificar se há registros
        if not registros:
            raise ValueError("Nenhum registro encontrado para exportação")
        
        # Definir o nome do arquivo de saída
        mes_ano_str = mes_ano if mes_ano else datetime.now().strftime("%Y-%m")
        nome_arquivo = f"controle_horas_{mes_ano_str}.xlsx"
        caminho_saida = os.path.join(os.path.dirname(os.path.dirname(__file__)), nome_arquivo)
        
        # Gerar o relatório personalizado
        return gerar_relatorio_mensal_personalizado(
            registros=registros,
            funcionarios=funcionarios,
            projetos=projetos,
            mes_ano=mes_ano_str,
            caminho_saida=caminho_saida
        )
    except Exception as e:
        print(f"Erro na adaptação da exportação: {e}")
        raise
